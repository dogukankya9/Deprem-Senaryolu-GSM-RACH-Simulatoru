import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import time
import csv
import os

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class RachSimulationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GSM Deprem Modu - RACH Çakışması Simülatörü")
        self.root.geometry("1240x790")
        self.root.configure(bg="#d9d9d9")
        self.root.resizable(True, True)
        self.root.minsize(760, 520)

        # -----------------------------
        # SİSTEM DURUM DEĞİŞKENLERİ
        # -----------------------------
        self.earthquake_mode = False
        self.earthquake_start_time = None

        # AYRI PAUSE / RESUME
        self.paused = False
        self.pause_start_time = None
        self.total_paused_duration = 0

        self.network_load = 25
        self.barring_level = 10
        self.collision_rate = 0
        self.total_attempts = 0
        self.success_count = 0
        self.blocked_count = 0
        self.collided_count = 0
        self.tick_count = 0

        # Grafik geçmiş verileri
        self.graph_ticks = []
        self.graph_network_load = []
        self.graph_barring_level = []
        self.graph_collision_rate = []
        self.max_graph_points = 60
        self.graph_visible = True

        # Manuel trafik butonları artık anlık yeni tur oluşturmaz.
        # Bu değer bir sonraki otomatik turda kullanılır ve sonra sıfırlanır.
        self.pending_traffic_boost = 0
        self.pending_traffic_label = None

        self.manual_seismic_value = 0.0
        self.use_manual_seismic = True
        self.last_seismic_value = 0.0

        self.user_classes = {
            "AFAD / Ambulans / İtfaiye": {"priority": 1, "color": "#b30000"},
            "Kamu / Kritik Altyapı": {"priority": 2, "color": "#cc5500"},
            "Normal Vatandaş": {"priority": 3, "color": "#004c99"},
            "Düşük Öncelik / Toplu Trafik": {"priority": 4, "color": "#666666"}
        }

        self.logs = []
        self.history_tables = {}
        self.history_stats = {}
        self.log_items = []
        self.selected_log_index = None
        self.selected_tick_no = None
        self.showing_total_status = True
        self.graph_window_size = 60
        self.graph_scroll_var = tk.IntVar(value=0)
        self.priority_stats_window = None
        self.priority_stats_cards = {}
        self.priority_stats_tree = None

        # -----------------------------
        # BACKOFF KUYRUĞU
        # -----------------------------
        # Her giriş: {"id", "class", "priority", "barring_prob",
        #              "attempt": kaçıncı deneme, "next_tick": hangi turda denenecek}
        self.backoff_queue = []
        self.backoff_count = 0
        self.timeout_count = 0
        self.BACKOFF_MAX_ATTEMPTS = 5
        self.BACKOFF_BASE = 2

        self.build_ui()
        self.resize_grip = ttk.Sizegrip(self.root)
        self.resize_grip.place(relx=1.0, rely=1.0, anchor='se')
        self.refresh_labels(self.last_seismic_value)
        self.refresh_button_states()
        self.update_loop()

    # =========================================================
    # ARAYÜZ
    # =========================================================
    def build_ui(self):
        title = tk.Label(
            self.root,
            text="GSM DEPREM MODU / RACH ÇAKIŞMASI SİMÜLATÖRÜ",
            font=("Courier New", 18, "bold"),
            bg="#a6a6a6",
            fg="black",
            relief="raised",
            bd=3,
            pady=8
        )
        title.pack(fill="x", padx=8, pady=8)

        main_frame = tk.Frame(self.root, bg="#d9d9d9")
        main_frame.pack(fill="both", expand=True, padx=8, pady=4)

        # SOL TARAF SCROLL
        left_outer = tk.Frame(main_frame, bg="#cfcfcf", relief="sunken", bd=2)
        left_outer.pack(side="left", fill="y", padx=(0, 6))

        self.left_canvas = tk.Canvas(left_outer, bg="#cfcfcf", highlightthickness=0, width=300)
        self.left_canvas.pack(side="left", fill="y", expand=False)

        left_scrollbar = tk.Scrollbar(left_outer, orient="vertical", command=self.left_canvas.yview)
        left_scrollbar.pack(side="right", fill="y")

        self.left_canvas.configure(yscrollcommand=left_scrollbar.set)

        left_panel = tk.Frame(self.left_canvas, bg="#cfcfcf")
        self.left_panel = left_panel

        self.left_canvas_window = self.left_canvas.create_window((0, 0), window=left_panel, anchor="nw")

        left_panel.bind("<Configure>", self.on_left_panel_configure)
        self.left_canvas.bind("<Configure>", self.on_left_canvas_configure)

        self.left_canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        self.left_canvas.bind_all("<Button-4>", self.on_mousewheel_linux)
        self.left_canvas.bind_all("<Button-5>", self.on_mousewheel_linux)

        # SAĞ PANEL
        right_panel = tk.Frame(main_frame, bg="#cfcfcf", relief="sunken", bd=2)
        right_panel.pack(side="right", fill="both", expand=True)

        info_title = tk.Label(
            left_panel, text="SİSTEM DURUMU",
            font=("Courier New", 13, "bold"),
            bg="#808080", fg="white", width=34
        )
        info_title.pack(fill="x", pady=4)

        self.lbl_earthquake = self.make_info_label(left_panel, "Deprem Modu: KAPALI")
        self.lbl_mode_time = self.make_info_label(left_panel, "Deprem Modu Süresi: 0 sn")
        self.lbl_pause = self.make_info_label(left_panel, "Simülasyon Durumu: ÇALIŞIYOR")
        self.lbl_seismic = self.make_info_label(left_panel, "Sismik Seviye: 0.00")
        self.lbl_network = self.make_info_label(left_panel, "Şebeke Yükü: %25")
        self.lbl_barring = self.make_info_label(left_panel, "Barring Seviyesi: %10")
        self.lbl_collision = self.make_info_label(left_panel, "RACH Çakışma: %0")
        self.lbl_total = self.make_info_label(left_panel, "Toplam Deneme: 0")
        self.lbl_success = self.make_info_label(left_panel, "Başarılı Erişim: 0")
        self.lbl_blocked = self.make_info_label(left_panel, "Barring Engeli: 0")
        self.lbl_collided = self.make_info_label(left_panel, "Çakışan Paket: 0")
        self.lbl_backoff = self.make_info_label(left_panel, "Backoff Kuyruğu: 0")
        self.lbl_timeout = self.make_info_label(left_panel, "Timeout (Max Deneme): 0")

        self.lbl_success_rate = tk.Label(
            left_panel,
            text="Toplam Başarı Oranı: %0",
            font=("Courier New", 11, "bold"),
            bg="#00c853",
            fg="white",
            anchor="center",
            relief="raised",
            bd=4,
            padx=6,
            pady=7,
            width=34
        )
        self.lbl_success_rate.pack(fill="x", padx=6, pady=(4, 2))

        ttk.Separator(left_panel, orient="horizontal").pack(fill="x", padx=6, pady=8)

        control_title = tk.Label(
            left_panel, text="KONTROLLER",
            font=("Courier New", 12, "bold"),
            bg="#808080", fg="white", width=34
        )
        control_title.pack(fill="x", pady=4)

        seismic_frame = tk.Frame(left_panel, bg="#cfcfcf")
        seismic_frame.pack(fill="x", padx=6, pady=4)

        tk.Label(
            seismic_frame,
            text="Elle Sismik Değer:",
            font=("Courier New", 10),
            bg="#cfcfcf"
        ).pack(anchor="w")

        self.seismic_scale = tk.Scale(
            seismic_frame, from_=0, to=20, resolution=0.1,
            orient="horizontal", length=220,
            font=("Courier New", 9),
            command=self.on_manual_seismic_change
        )
        self.seismic_scale.set(0)
        self.seismic_scale.pack()

        self.chk_manual_var = tk.IntVar(value=1)
        chk_manual = tk.Checkbutton(
            seismic_frame,
            text="Elle sismik veri kullan",
            variable=self.chk_manual_var,
            font=("Courier New", 10),
            bg="#cfcfcf",
            command=self.toggle_manual_seismic
        )
        chk_manual.pack(anchor="w", pady=3)

        btn_frame = tk.Frame(left_panel, bg="#cfcfcf")
        btn_frame.pack(fill="x", padx=6, pady=6)

        # TEK DEPREM MODU BUTONU
        self.btn_trigger = tk.Button(
            btn_frame,
            text="DEPREM MODUNU TETİKLE",
            font=("Courier New", 10, "bold"),
            bg="#ffcc00",
            activebackground="#e6b800",
            width=28,
            relief="raised",
            bd=3,
            command=self.toggle_earthquake_mode
        )
        self.btn_trigger.pack(pady=3)

        # HER ZAMAN AKTİF PAUSE / RESUME
        self.btn_pause_resume = tk.Button(
            btn_frame,
            text="SİMÜLASYONU DURDUR",
            font=("Courier New", 10, "bold"),
            bg="#cccc00",
            activebackground="#b3b300",
            width=28,
            relief="raised",
            bd=3,
            state="normal",
            command=self.toggle_pause_resume
        )
        self.btn_pause_resume.pack(pady=3)

        self.btn_normal = tk.Button(
            btn_frame,
            text="NORMAL TRAFİK ÜRET",
            font=("Courier New", 10, "bold"),
            bg="#99ccff",
            activebackground="#7fbfff",
            width=28,
            relief="raised",
            bd=3,
            command=lambda: self.manual_generate_traffic(20, self.btn_normal)
        )
        self.btn_normal.pack(pady=3)

        self.btn_heavy = tk.Button(
            btn_frame,
            text="YOĞUN TRAFİK ÜRET",
            font=("Courier New", 10, "bold"),
            bg="#ff9966",
            activebackground="#ff8552",
            width=28,
            relief="raised",
            bd=3,
            command=lambda: self.manual_generate_traffic(60, self.btn_heavy)
        )
        self.btn_heavy.pack(pady=3)

        self.btn_graph_toggle = tk.Button(
            btn_frame,
            text="GRAFİĞİ GİZLE",
            font=("Courier New", 10, "bold"),
            bg="#b3e5fc",
            activebackground="#81d4fa",
            width=28,
            relief="raised",
            bd=3,
            command=self.toggle_graph_visibility
        )
        self.btn_graph_toggle.pack(pady=3)


        self.btn_total_status = tk.Button(
            btn_frame,
            text="GENEL TOPLAMI GÖSTER",
            font=("Courier New", 10, "bold"),
            bg="#a5d6a7",
            activebackground="#81c784",
            width=28,
            relief="raised",
            bd=3,
            command=self.show_total_status
        )
        self.btn_total_status.pack(pady=3)

        self.btn_priority_stats = tk.Button(
            btn_frame,
            text="ÖNCELİK İSTATİSTİKLERİ",
            font=("Courier New", 10, "bold"),
            bg="#d1c4e9",
            activebackground="#b39ddb",
            width=28,
            relief="raised",
            bd=3,
            command=self.open_priority_stats_window
        )
        self.btn_priority_stats.pack(pady=3)

        self.btn_reset = tk.Button(
            btn_frame,
            text="SAYAÇLARI SIFIRLA",
            font=("Courier New", 10, "bold"),
            bg="#dddddd",
            activebackground="#c8c8c8",
            width=28,
            relief="raised",
            bd=3,
            command=self.reset_counters
        )
        self.btn_reset.pack(pady=3)

        ttk.Separator(btn_frame, orient="horizontal").pack(fill="x", pady=6)

        export_title = tk.Label(
            btn_frame, text="DIŞA AKTAR",
            font=("Courier New", 10, "bold"),
            bg="#808080", fg="white", width=28
        )
        export_title.pack(pady=(0, 3))

        self.btn_export_csv = tk.Button(
            btn_frame,
            text="CSV OLARAK KAYDET",
            font=("Courier New", 10, "bold"),
            bg="#e8f5e9",
            activebackground="#c8e6c9",
            width=28,
            relief="raised",
            bd=3,
            command=self.export_csv
        )
        self.btn_export_csv.pack(pady=3)

        excel_bg = "#e3f2fd" if EXCEL_AVAILABLE else "#eeeeee"
        excel_text = "EXCEL OLARAK KAYDET" if EXCEL_AVAILABLE else "EXCEL (openpyxl yok)"
        self.btn_export_excel = tk.Button(
            btn_frame,
            text=excel_text,
            font=("Courier New", 10, "bold"),
            bg=excel_bg,
            activebackground="#bbdefb",
            width=28,
            relief="raised",
            bd=3,
            state="normal" if EXCEL_AVAILABLE else "disabled",
            command=self.export_excel
        )
        self.btn_export_excel.pack(pady=3)

        ttk.Separator(left_panel, orient="horizontal").pack(fill="x", padx=6, pady=8)

        legend_title = tk.Label(
            left_panel, text="ÖNCELİK SINIFLARI",
            font=("Courier New", 12, "bold"),
            bg="#808080", fg="white", width=34
        )
        legend_title.pack(fill="x", pady=4)

        for name, data in self.user_classes.items():
            lbl = tk.Label(
                left_panel,
                text=f"P{data['priority']} - {name}",
                font=("Courier New", 10, "bold"),
                bg=data["color"],
                fg="white",
                anchor="w",
                padx=6
            )
            lbl.pack(fill="x", padx=6, pady=2)


        tk.Label(left_panel, text="", bg="#cfcfcf", height=1).pack()

        table_title = tk.Label(
            right_panel, text="ERİŞİM DENEME TABLOSU",
            font=("Courier New", 13, "bold"),
            bg="#808080", fg="white"
        )
        table_title.pack(fill="x", pady=4)

        self.lbl_table_info = tk.Label(
            right_panel,
            text="Gösterilen Tablo: Son Tur",
            font=("Courier New", 10, "bold"),
            bg="#e6e6e6",
            anchor="w",
            relief="groove",
            bd=2,
            padx=6,
            pady=4
        )
        self.lbl_table_info.pack(fill="x", padx=8, pady=(2, 4))

        self.lbl_selected_log = tk.Label(
            right_panel,
            text="Seçilen Log: Yok",
            font=("Courier New", 10, "bold"),
            bg="#fff2cc",
            anchor="w",
            relief="groove",
            bd=2,
            padx=6,
            pady=4
        )
        self.lbl_selected_log.pack(fill="x", padx=8, pady=(0, 6))

        self.lbl_slot_capacity = tk.Label(
            right_panel,
            text="Anlık RACH Slot Kapasitesi: 0",
            font=("Courier New", 10, "bold"),
            bg="#d9ead3",
            anchor="w",
            relief="groove",
            bd=2,
            padx=6,
            pady=4
        )
        self.lbl_slot_capacity.pack(fill="x", padx=8, pady=(0, 6))

        columns = ("id", "sinif", "oncelik", "istek", "barring", "backoff", "sonuc")
        self.tree = ttk.Treeview(right_panel, columns=columns, show="headings", height=14)

        self.tree.heading("id", text="ID")
        self.tree.heading("sinif", text="Kullanıcı Sınıfı")
        self.tree.heading("oncelik", text="Öncelik")
        self.tree.heading("istek", text="İstek Sayısı")
        self.tree.heading("barring", text="Barring")
        self.tree.heading("backoff", text="Deneme")
        self.tree.heading("sonuc", text="Sonuç")

        self.tree.column("id", width=80, anchor="center")
        self.tree.column("sinif", width=220, anchor="center")
        self.tree.column("oncelik", width=70, anchor="center")
        self.tree.column("istek", width=90, anchor="center")
        self.tree.column("barring", width=90, anchor="center")
        self.tree.column("backoff", width=80, anchor="center")
        self.tree.column("sonuc", width=220, anchor="center")

        self.tree.pack(fill="x", padx=8, pady=6)

        # Grafik ve Sistem Olay Kayıtları bölümleri artık mouse ile sürüklenebilir.
        self.right_vertical_pane = tk.PanedWindow(
            right_panel,
            orient="vertical",
            sashwidth=8,
            sashrelief="raised",
            bg="#b0b0b0",
            bd=0
        )
        self.right_vertical_pane.pack(fill="both", expand=True, padx=8, pady=(8, 6))

        self.graph_area = tk.Frame(self.right_vertical_pane, bg="#cfcfcf", relief="groove", bd=2)
        self.log_area = tk.Frame(self.right_vertical_pane, bg="#cfcfcf", relief="groove", bd=2)

        self.right_vertical_pane.add(self.graph_area, minsize=120, stretch="always")
        self.right_vertical_pane.add(self.log_area, minsize=120, stretch="always")

        self.graph_title = tk.Label(
            self.graph_area, text="CANLI PERFORMANS GRAFİĞİ",
            font=("Courier New", 13, "bold"),
            bg="#808080", fg="white"
        )
        self.graph_title.pack(fill="x", pady=(0, 4))

        self.figure = Figure(figsize=(7.8, 2.4), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("Şebeke Yükü / Barring / RACH Çakışma")
        self.ax.set_xlabel("Tur")
        self.ax.set_ylabel("Yüzde (%)")
        self.ax.set_ylim(0, 100)
        self.ax.grid(True, linestyle="--", alpha=0.4)

        self.graph_canvas = FigureCanvasTkAgg(self.figure, master=self.graph_area)
        self.graph_canvas.get_tk_widget().pack(fill="both", expand=True, padx=6, pady=(4, 2))

        self.graph_scroll = tk.Scale(
            self.graph_area,
            from_=0,
            to=0,
            orient="horizontal",
            variable=self.graph_scroll_var,
            showvalue=False,
            command=self.on_graph_scroll,
            bg="#cfcfcf",
            length=500
        )
        self.graph_scroll.pack(fill="x", padx=6, pady=(0, 4))

        self.log_title = tk.Label(
            self.log_area, text="SİSTEM OLAY KAYITLARI",
            font=("Courier New", 13, "bold"),
            bg="#808080", fg="white"
        )
        self.log_title.pack(fill="x", pady=(0, 4))

        log_frame = tk.Frame(self.log_area, bg="#cfcfcf")
        log_frame.pack(fill="both", expand=True, padx=6, pady=6)

        self.log_listbox = tk.Listbox(
            log_frame,
            font=("Courier New", 10),
            bg="black",
            fg="#00ff00",
            selectbackground="#1e5eff",
            selectforeground="white",
            activestyle="dotbox",
            exportselection=False
        )
        self.log_listbox.pack(side="left", fill="both", expand=True)

        log_scroll = tk.Scrollbar(log_frame, command=self.log_listbox.yview)
        log_scroll.pack(side="right", fill="y")
        self.log_listbox.config(yscrollcommand=log_scroll.set)

        self.log_listbox.bind("<<ListboxSelect>>", self.on_log_select)

        # Boş alana tıklanınca seçili turdan genel toplam görünümüne dön
        self.root.bind("<Button-1>", self.on_global_click, add="+")
        right_panel.bind("<Button-1>", self.on_global_click, add="+")
        self.tree.bind("<Button-1>", self.on_global_click, add="+")

        self.add_log("Sistem başlatıldı.")
        self.add_log("Normal çalışma modu aktif.")
        self.add_log("Elle sismik veri modu varsayılan olarak açık.")
        self.add_log("Log satırına tıklayarak ilgili turun tablosu görüntülenebilir.")

    # =========================================================
    # SCROLL
    # =========================================================
    def on_left_panel_configure(self, event):
        self.left_canvas.configure(scrollregion=self.left_canvas.bbox("all"))

    def on_left_canvas_configure(self, event):
        self.left_canvas.itemconfig(self.left_canvas_window, width=event.width)

    def on_mousewheel(self, event):
        try:
            widget = self.root.winfo_containing(event.x_root, event.y_root)
            if widget and self.is_widget_inside_left_panel(widget):
                self.left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass

    def on_mousewheel_linux(self, event):
        try:
            widget = self.root.winfo_containing(event.x_root, event.y_root)
            if widget and self.is_widget_inside_left_panel(widget):
                if event.num == 4:
                    self.left_canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    self.left_canvas.yview_scroll(1, "units")
        except Exception:
            pass

    def is_widget_inside_left_panel(self, widget):
        parent = widget
        while parent is not None:
            if parent == self.left_panel:
                return True
            parent = parent.master
        return False

    # =========================================================
    # UI
    # =========================================================
    def make_info_label(self, parent, text):
        lbl = tk.Label(
            parent,
            text=text,
            font=("Courier New", 10, "bold"),
            bg="#e6e6e6",
            anchor="w",
            relief="groove",
            bd=2,
            padx=6,
            pady=4,
            width=34
        )
        lbl.pack(fill="x", padx=6, pady=2)
        return lbl

    def set_button_pressed_effect(self, button):
        button.config(relief="sunken", bd=5)
        self.root.after(180, lambda: button.config(relief="raised", bd=3))

    def refresh_button_states(self):
        # Tek deprem modu butonu
        if self.earthquake_mode:
            self.btn_trigger.config(
                state="normal",
                bg="#ff6666",
                activebackground="#e05555",
                text="DEPREM MODUNU DURDUR"
            )
        else:
            self.btn_trigger.config(
                state="normal",
                bg="#ffcc00",
                activebackground="#e6b800",
                text="DEPREM MODUNU TETİKLE"
            )

        # Pause / Resume HER ZAMAN aktif
        if self.paused:
            self.btn_pause_resume.config(
                state="normal",
                text="DEVAM ET",
                bg="#66cc66",
                activebackground="#52b352"
            )
            self.btn_normal.config(state="disabled", disabledforeground="#666666", bg="#cfd8dc")
            self.btn_heavy.config(state="disabled", disabledforeground="#666666", bg="#d7ccc8")
        else:
            self.btn_pause_resume.config(
                state="normal",
                text="SİMÜLASYONU DURDUR",
                bg="#cccc00",
                activebackground="#b3b300"
            )
            self.btn_normal.config(state="normal", bg="#99ccff")
            self.btn_heavy.config(state="normal", bg="#ff9966")

    # =========================================================
    # LOG
    # =========================================================
    def add_log(self, message, tick_no=None):
        timestamp = time.strftime("%H:%M:%S")
        line = f"[{timestamp}] {message}"

        self.logs.append(line)
        self.log_items.append((line, tick_no))

        if len(self.log_items) > 400:
            self.log_items.pop(0)

        self.refresh_log_listbox()

    def refresh_log_listbox(self):
        old_first, old_last = self.log_listbox.yview() if hasattr(self, "log_listbox") else (0.0, 1.0)
        should_follow_live = (not self.paused) and (old_last >= 0.98)

        self.log_listbox.delete(0, tk.END)
        for text, _ in self.log_items:
            self.log_listbox.insert(tk.END, text)

        if should_follow_live:
            self.log_listbox.see(tk.END)
            self.log_listbox.yview_moveto(1.0)
        elif self.paused:
            self.log_listbox.yview_moveto(old_first)

    def clear_logs(self):
        self.logs.clear()
        self.log_items.clear()
        self.selected_log_index = None
        self.selected_tick_no = None
        self.log_listbox.delete(0, tk.END)
        self.lbl_selected_log.config(text="Seçilen Log: Yok")

    def on_log_select(self, event):
        selection = self.log_listbox.curselection()

        if not selection:
            self.selected_log_index = None
            self.selected_tick_no = None
            self.lbl_selected_log.config(text="Seçilen Log: Yok")
            return

        index = selection[0]
        self.selected_log_index = index

        text, tick_no = self.log_items[index]

        if tick_no is None:
            self.selected_tick_no = None
            self.lbl_selected_log.config(text="Seçilen Log: Genel sistem kaydı")
            self.lbl_table_info.config(text="Gösterilen Tablo: Bu log için kayıtlı tur yok")
            self.refresh_labels(self.last_seismic_value)
        else:
            self.selected_tick_no = tick_no
            self.showing_total_status = False
            self.lbl_selected_log.config(text=f"Seçilen Log: TUR {tick_no}")

            if tick_no in self.history_tables:
                rows = self.history_tables[tick_no]
                self.show_table_rows(rows)
                self.lbl_table_info.config(text=f"Gösterilen Tablo: Tur {tick_no}")

                tur_total = len(rows)
                tur_success = sum(1 for x in rows if x["result"] == "ERİŞİM BAŞARILI")
                tur_blocked = sum(1 for x in rows if x["result"] == "BARRING ENGELİ")
                tur_collided = sum(1 for x in rows if x["result"] == "RACH ÇAKIŞMASI")
                tur_rate = int((tur_success / tur_total) * 100) if tur_total > 0 else 0

                self.lbl_total.config(text=f"Tur Deneme: {tur_total}")
                self.lbl_success.config(text=f"Tur Başarılı: {tur_success}")
                self.lbl_blocked.config(text=f"Tur Barring Engeli: {tur_blocked}")
                self.lbl_collided.config(text=f"Tur Çakışan Paket: {tur_collided}")
                self.lbl_success_rate.config(
                    text=f"Tur Başarı Oranı: %{tur_rate}",
                    bg="#66cc66",
                    fg="black"
                )
                self.show_tick_status(tick_no)
            else:
                self.lbl_table_info.config(text=f"Gösterilen Tablo: Tur {tick_no} bulunamadı")

        if not self.paused:
            self.root.after(10, lambda: self.log_listbox.yview_moveto(1.0))


    def on_global_click(self, event):
        """Log listesinin boş olmayan satırı dışında bir yere tıklanınca genel toplam görünümüne döner."""
        widget = event.widget

        # Log listesinde gerçek bir satıra tıklanıyorsa tur seçimi çalışsın, genel görünüme dönmesin.
        if widget == self.log_listbox:
            try:
                nearest = self.log_listbox.nearest(event.y)
                bbox = self.log_listbox.bbox(nearest)
                if bbox is not None:
                    x, y, w, h = bbox
                    if y <= event.y <= y + h:
                        return
            except Exception:
                pass

        self.clear_log_selection_and_show_total()

    def clear_log_selection_and_show_total(self):
        self.selected_log_index = None
        self.selected_tick_no = None

        try:
            self.log_listbox.selection_clear(0, tk.END)
        except Exception:
            pass

        # Genel toplam değerlerine dön
        self.refresh_labels(self.last_seismic_value)
        self.lbl_selected_log.config(text="Seçilen Log: Yok / Genel toplam gösteriliyor")
        self.lbl_table_info.config(text=f"Gösterilen Tablo: Son Tur (Genel Sistem Durumu)")

    # =========================================================
    # TABLO
    # =========================================================
    def show_table_rows(self, rows):
        self.tree.delete(*self.tree.get_children())

        # Renk etiketleri
        self.tree.tag_configure("success",  background="#e8f5e9", foreground="#1b5e20")
        self.tree.tag_configure("barred",   background="#fff3e0", foreground="#bf360c")
        self.tree.tag_configure("collided", background="#fce4ec", foreground="#880e4f")
        self.tree.tag_configure("backoff",  background="#fffde7", foreground="#f57f17")
        self.tree.tag_configure("timeout",  background="#eeeeee", foreground="#616161")

        for item in rows:
            result = item.get("result", "")
            attempt = item.get("attempt", 1)

            if result == "ERİŞİM BAŞARILI":
                tag = "success"
            elif result == "BARRING ENGELİ":
                tag = "barred"
            elif result.startswith("RACH ÇAKIŞMASI"):
                tag = "collided"
            elif result == "BACKOFF BEKLİYOR":
                tag = "backoff"
            elif result == "TIMEOUT":
                tag = "timeout"
            else:
                tag = ""

            self.tree.insert(
                "",
                "end",
                values=(
                    item["id"],
                    item["class"],
                    f"P{item['priority']}",
                    1,
                    f"%{item['barring_prob']}",
                    f"{attempt}/{self.BACKOFF_MAX_ATTEMPTS}",
                    result
                ),
                tags=(tag,)
            )

    # =========================================================
    # GRAFİK
    # =========================================================
    def redraw_graph(self):
        self.ax.clear()

        if not self.graph_ticks:
            self.ax.set_title("Şebeke Yükü / Barring / RACH Çakışma")
            self.ax.set_xlabel("Tur")
            self.ax.set_ylabel("Yüzde (%)")
            self.ax.set_ylim(0, 100)
            self.ax.grid(True, linestyle="--", alpha=0.4)
            if self.graph_visible:
                self.graph_canvas.draw_idle()
            return

        max_start = max(0, len(self.graph_ticks) - self.graph_window_size)
        start_index = min(self.graph_scroll_var.get(), max_start)
        end_index = min(len(self.graph_ticks), start_index + self.graph_window_size)

        ticks = self.graph_ticks[start_index:end_index]
        network = self.graph_network_load[start_index:end_index]
        barring = self.graph_barring_level[start_index:end_index]
        collision = self.graph_collision_rate[start_index:end_index]

        self.ax.plot(ticks, network, label="Şebeke Yükü (%)")
        self.ax.plot(ticks, barring, label="Barring (%)")
        self.ax.plot(ticks, collision, label="RACH Çakışma (%)")
        self.ax.set_title("Şebeke Yükü / Barring / RACH Çakışma")
        self.ax.set_xlabel("Tur")
        self.ax.set_ylabel("Yüzde (%)")
        self.ax.set_ylim(0, 100)
        self.ax.grid(True, linestyle="--", alpha=0.4)
        self.ax.legend(loc="upper left", fontsize=8)
        self.figure.tight_layout()


        if self.graph_visible:
            self.graph_canvas.draw_idle()

    def on_graph_scroll(self, value=None):
        self.redraw_graph()

    def update_graph(self):
        self.graph_ticks.append(self.tick_count)
        self.graph_network_load.append(self.network_load)
        self.graph_barring_level.append(self.barring_level)
        self.graph_collision_rate.append(self.collision_rate)

        max_start = max(0, len(self.graph_ticks) - self.graph_window_size)
        self.graph_scroll.config(to=max_start)

        if self.graph_scroll_var.get() >= max(0, max_start - 1):
            self.graph_scroll_var.set(max_start)

        self.redraw_graph()

    def toggle_graph_visibility(self):
        self.set_button_pressed_effect(self.btn_graph_toggle)

        if self.graph_visible:
            try:
                self.right_vertical_pane.forget(self.graph_area)
            except Exception:
                pass

            self.graph_visible = False
            self.btn_graph_toggle.config(
                text="GRAFİĞİ GÖSTER",
                bg="#c8e6c9",
                activebackground="#a5d6a7"
            )
            self.add_log("Canlı performans grafiği gizlendi.")
        else:
            # Her iki bölmeyi de çıkarıp doğru sırayla tekrar ekle:
            # grafik üstte, log altta
            for pane in (self.graph_area, self.log_area):
                try:
                    self.right_vertical_pane.forget(pane)
                except Exception:
                    pass
            self.right_vertical_pane.add(self.graph_area, minsize=120, stretch="always")
            self.right_vertical_pane.add(self.log_area, minsize=120, stretch="always")

            self.graph_visible = True
            self.btn_graph_toggle.config(
                text="GRAFİĞİ GİZLE",
                bg="#b3e5fc",
                activebackground="#81d4fa"
            )
            self.add_log("Canlı performans grafiği gösterildi.")
            self.graph_canvas.draw_idle()

    # =========================================================
    # YARDIMCI
    # =========================================================
    def on_manual_seismic_change(self, value):
        self.manual_seismic_value = float(value)
        self.last_seismic_value = self.manual_seismic_value
        self.refresh_labels(self.last_seismic_value)

    def toggle_manual_seismic(self):
        self.use_manual_seismic = bool(self.chk_manual_var.get())
        if self.use_manual_seismic:
            self.add_log("Elle sismik veri modu açıldı.")
        else:
            self.add_log("Otomatik sismik veri modu açıldı.")

    def manual_generate_traffic(self, boost, button):
        if self.paused:
            return

        self.set_button_pressed_effect(button)

        # HATA DÜZELTME:
        # Eski sürümde bu buton doğrudan generate_traffic() çağırıyordu.
        # Bu yüzden aynı saniye içinde ekstra tur oluşuyordu.
        # Yeni mantıkta buton sadece bir sonraki otomatik tura trafik yükü ekler.
        self.pending_traffic_boost += boost
        self.pending_traffic_label = "YOĞUN" if boost >= 60 else "NORMAL"

        self.add_log(
            f"{self.pending_traffic_label} trafik isteği alındı. Yük bir sonraki turda uygulanacak."
        )
        self.refresh_labels(self.last_seismic_value)

    # =========================================================
    # PAUSE / RESUME
    # =========================================================
    def toggle_pause_resume(self):
        self.set_button_pressed_effect(self.btn_pause_resume)

        if not self.paused:
            self.pause_simulation()
        else:
            self.resume_simulation()

    def pause_simulation(self):
        if not self.paused:
            self.paused = True
            self.pause_start_time = time.time()
            self.add_log("Simülasyon DURDURULDU. Tüm değerler donduruldu.")
            self.refresh_labels(self.last_seismic_value)
            self.refresh_button_states()

    def resume_simulation(self):
        if self.paused:
            paused_duration = time.time() - self.pause_start_time
            self.total_paused_duration += paused_duration
            self.paused = False
            self.pause_start_time = None
            self.add_log("Simülasyon tekrar başlatıldı.")
            self.refresh_labels(self.last_seismic_value)
            self.refresh_button_states()

    # =========================================================
    # DEPREM MODU
    # =========================================================
    def toggle_earthquake_mode(self):
        self.set_button_pressed_effect(self.btn_trigger)

        if not self.earthquake_mode:
            self.earthquake_mode = True
            self.earthquake_start_time = time.time()

            self.add_log("Deprem modu MANUEL olarak aktifleştirildi.")
        else:
            self.earthquake_mode = False
            self.earthquake_start_time = None
            self.barring_level = 10

            self.add_log("Deprem modu kullanıcı tarafından durduruldu.")

        self.refresh_labels(self.last_seismic_value)
        self.refresh_button_states()

    def reset_counters(self):
        self.set_button_pressed_effect(self.btn_reset)

        self.total_attempts = 0
        self.success_count = 0
        self.blocked_count = 0
        self.collided_count = 0
        self.collision_rate = 0
        self.tick_count = 0
        self.pending_traffic_boost = 0
        self.pending_traffic_label = None
        self.backoff_queue.clear()
        self.backoff_count = 0
        self.timeout_count = 0

        self.tree.delete(*self.tree.get_children())
        self.history_tables.clear()
        self.history_stats.clear()
        self.graph_ticks.clear()
        self.graph_network_load.clear()
        self.graph_barring_level.clear()
        self.graph_collision_rate.clear()
        self.graph_scroll_var.set(0)
        self.graph_scroll.config(to=0)
        self.ax.clear()
        self.ax.set_title("Şebeke Yükü / Barring / RACH Çakışma")
        self.ax.set_xlabel("Tur")
        self.ax.set_ylabel("Yüzde (%)")
        self.ax.set_ylim(0, 100)
        self.ax.grid(True, linestyle="--", alpha=0.4)
        self.graph_canvas.draw_idle()
        self.lbl_table_info.config(text="Gösterilen Tablo: Son Tur")
        if hasattr(self, "lbl_slot_capacity"):
            self.lbl_slot_capacity.config(text=f"Anlık RACH Slot Kapasitesi: {self.calculate_available_slots()} Slot")

        self.clear_logs()
        self.add_log("Sistem olay kayıtları ve sayaçlar sıfırlandı.")

        self.refresh_labels(self.last_seismic_value)
        self.refresh_priority_stats_window()


    # =========================================================
    # DIŞA AKTARIM
    # =========================================================
    def _build_export_rows(self):
        """Tüm history_tables verisini düz satır listesine çevirir."""
        rows = []
        for tick_no in sorted(self.history_tables.keys()):
            stats = self.history_stats.get(tick_no, {})
            for item in self.history_tables[tick_no]:
                rows.append({
                    "Tur":             tick_no,
                    "Sismik Değer":    stats.get("seismic_value", ""),
                    "Şebeke Yükü (%)": stats.get("network_load", ""),
                    "Barring (%)":     stats.get("barring_level", ""),
                    "Slot Kapasitesi": stats.get("available_slots", ""),
                    "Paket ID":        item.get("id", ""),
                    "Kullanıcı Sınıfı":item.get("class", ""),
                    "Öncelik":         item.get("priority", ""),
                    "Barring Olasılığı (%)": item.get("barring_prob", ""),
                    "Deneme No":       item.get("attempt", 1),
                    "Sonuç":           item.get("result", ""),
                })
        return rows

    def export_csv(self):
        if not self.history_tables:
            messagebox.showwarning("Dışa Aktarım", "Henüz kayıtlı tur verisi yok.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Dosyası", "*.csv")],
            initialfile=f"rach_sim_{time.strftime('%Y%m%d_%H%M%S')}.csv",
            title="CSV Olarak Kaydet"
        )
        if not filepath:
            return

        rows = self._build_export_rows()
        try:
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)
            self.add_log(f"CSV dışa aktarıldı: {os.path.basename(filepath)} ({len(rows)} satır)")
            messagebox.showinfo("Başarılı", f"{len(rows)} satır CSV olarak kaydedildi.\n{filepath}")
        except Exception as e:
            messagebox.showerror("Hata", f"CSV kaydedilemedi:\n{e}")

    def export_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil.\npip install openpyxl")
            return
        if not self.history_tables:
            messagebox.showwarning("Dışa Aktarım", "Henüz kayıtlı tur verisi yok.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile=f"rach_sim_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
            title="Excel Olarak Kaydet"
        )
        if not filepath:
            return

        rows = self._build_export_rows()
        try:
            wb = openpyxl.Workbook()

            # ---- Sayfa 1: Paket Detayları ----
            ws1 = wb.active
            ws1.title = "Paket Detayları"

            header_fill  = PatternFill("solid", fgColor="2C3E50")
            header_font  = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border  = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            result_colors = {
                "ERİŞİM BAŞARILI": "D5E8D4",
                "BARRING ENGELİ":  "FFE6CC",
                "TIMEOUT":         "E0E0E0",
            }

            headers = list(rows[0].keys())
            for col_idx, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=h)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = thin_border

            for row_idx, row in enumerate(rows, 2):
                result = row.get("Sonuç", "")
                row_color = None
                for key, color in result_colors.items():
                    if result == key:
                        row_color = color
                        break
                if row_color is None and result.startswith("BACKOFF"):
                    row_color = "FFF2CC"

                for col_idx, key in enumerate(headers, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=row[key])
                    cell.alignment = center_align
                    cell.border    = thin_border
                    if row_color:
                        cell.fill = PatternFill("solid", fgColor=row_color)

            # Sütun genişlikleri
            col_widths = [6, 12, 14, 12, 14, 14, 28, 10, 20, 10, 24]
            for i, w in enumerate(col_widths, 1):
                ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            ws1.freeze_panes = "A2"

            # ---- Sayfa 2: Tur Özeti ----
            ws2 = wb.create_sheet("Tur Özeti")
            summary_headers = [
                "Tur", "Sismik", "Şebeke Yükü (%)", "Barring (%)",
                "Slot Kap.", "Toplam", "Başarılı", "Engelli",
                "Backoff/Timeout", "Başarı (%)", "Çakışma (%)"
            ]
            for col_idx, h in enumerate(summary_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = thin_border

            for row_idx, tick_no in enumerate(sorted(self.history_stats.keys()), 2):
                s = self.history_stats[tick_no]
                vals = [
                    tick_no,
                    s.get("seismic_value", 0),
                    s.get("network_load", 0),
                    s.get("barring_level", 0),
                    s.get("available_slots", 0),
                    s.get("total", 0),
                    s.get("success", 0),
                    s.get("blocked", 0),
                    s.get("collided", 0),
                    s.get("success_rate", 0),
                    s.get("collision_rate", 0),
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=v)
                    cell.alignment = center_align
                    cell.border    = thin_border

            for i, w in enumerate([6, 8, 16, 12, 10, 8, 10, 10, 16, 12, 12], 1):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws2.freeze_panes = "A2"

            wb.save(filepath)
            self.add_log(f"Excel dışa aktarıldı: {os.path.basename(filepath)} ({len(rows)} satır, {len(self.history_stats)} tur)")
            messagebox.showinfo("Başarılı", f"{len(rows)} satır Excel olarak kaydedildi.\n{filepath}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel kaydedilemedi:\n{e}")

    def show_total_summary(self):
        """show_total_status ile aynı işlevi görür; geriye dönük uyumluluk için korunmuştur."""
        self.show_total_status()

    # =========================================================
    # SİSMİK
    # =========================================================
    def get_seismic_value(self):
        if self.use_manual_seismic:
            return self.manual_seismic_value

        base = random.uniform(0.2, 3.5)
        if random.random() < 0.04:
            base += random.uniform(8, 15)

        return round(base, 2)

    def check_earthquake_trigger(self, seismic_value):
        threshold = 8.5
        if seismic_value >= threshold and not self.earthquake_mode:
            self.earthquake_mode = True
            self.earthquake_start_time = time.time()
            self.add_log(f"Sismik eşik aşıldı ({seismic_value}). Deprem modu aktif edildi.")
            self.refresh_button_states()

    def get_earthquake_mode_duration(self):
        if self.earthquake_mode and self.earthquake_start_time:
            if self.paused and self.pause_start_time is not None:
                return int(
                    self.pause_start_time
                    - self.earthquake_start_time
                    - self.total_paused_duration
                )
            else:
                return int(
                    time.time()
                    - self.earthquake_start_time
                    - self.total_paused_duration
                )
        return 0

    # =========================================================
    # ŞEBEKE
    # =========================================================
    def update_network_conditions(self):
        if not self.earthquake_mode:
            self.network_load += random.randint(-4, 4)
            self.network_load = max(10, min(55, self.network_load))
        else:
            self.network_load += random.randint(-3, 8)
            self.network_load = max(40, min(100, self.network_load))

        if self.earthquake_mode:
            if self.network_load >= 90:
                self.barring_level = 70
            elif self.network_load >= 80:
                self.barring_level = 55
            elif self.network_load >= 70:
                self.barring_level = 40
            elif self.network_load >= 60:
                self.barring_level = 25
            else:
                self.barring_level = 15
        else:
            # Agresif-adaptif barring mantığı:
            # %0–40 yük  -> barring yok
            # %40–55 yük -> hafif barring
            # %55–70 yük -> kontrollü barring
            # %70–85 yük -> güçlü barring
            # %85+ yük   -> çok güçlü barring
            if self.network_load >= 85:
                self.barring_level = 60
            elif self.network_load >= 70:
                self.barring_level = 35
            elif self.network_load >= 55:
                self.barring_level = 18
            elif self.network_load >= 40:
                self.barring_level = 8
            else:
                self.barring_level = 0

    def priority_barring_probability(self, priority):
        # Şebeke boşsa barring uygulanmaz
        if self.barring_level <= 0:
            return 0

        if priority == 1:
            return max(0, self.barring_level - 45)
        elif priority == 2:
            return max(0, self.barring_level - 20)
        elif priority == 3:
            return self.barring_level
        else:
            return min(95, self.barring_level + 15)

    def calculate_available_slots(self):
        if not self.earthquake_mode:
            if self.network_load < 40:
                return 12
            elif self.network_load < 55:
                return 10
            elif self.network_load < 70:
                return 8
            elif self.network_load < 85:
                return 5
            else:
                return 3
        else:
            if self.network_load < 60:
                return 6
            elif self.network_load < 80:
                return 4
            else:
                return 3

    # =========================================================
    # TRAFİK
    # =========================================================
    def generate_traffic(self, manual_boost=0):
        if self.paused:
            return

        self.tick_count += 1

        # ---- Yeni kullanıcılar oluştur ----
        if self.earthquake_mode:
            user_count = random.randint(8, 18) + manual_boost // 4
        else:
            user_count = random.randint(3, 8) + manual_boost // 5

        class_names = list(self.user_classes.keys())
        weighted_classes_normal = [
            class_names[0], class_names[1],
            class_names[2], class_names[2], class_names[2],
            class_names[3]
        ]
        weighted_classes_quake = [
            class_names[0], class_names[0], class_names[1],
            class_names[2], class_names[2], class_names[2], class_names[2],
            class_names[3]
        ]
        pool = weighted_classes_quake if self.earthquake_mode else weighted_classes_normal

        new_attempts = []
        for i in range(user_count):
            chosen = random.choice(pool)
            priority = self.user_classes[chosen]["priority"]
            barring_prob = self.priority_barring_probability(priority)
            new_attempts.append({
                "id": f"{self.tick_count}-{i+1}",
                "class": chosen,
                "priority": priority,
                "barring_prob": barring_prob,
                "barred": False,
                "attempt": 1,
                "result": ""
            })

        # ---- Backoff kuyruğundaki paketleri bu tura dahil et ----
        due_from_backoff = []
        still_waiting = []
        for pkt in self.backoff_queue:
            if pkt["next_tick"] <= self.tick_count:
                due_from_backoff.append(pkt)
            else:
                still_waiting.append(pkt)
        self.backoff_queue = still_waiting

        # Bu turda işlenecek tüm paketler = yeni + süresi dolan backoff
        attempts_this_round = new_attempts + due_from_backoff

        # ---- Barring uygula (sadece yeni paketlere) ----
        for item in new_attempts:
            is_barred = random.randint(1, 100) <= item["barring_prob"]
            item["barred"] = is_barred

        non_barred = [a for a in attempts_this_round if not a.get("barred", False)]
        barred     = [a for a in attempts_this_round if a.get("barred", False)]

        available_slots = self.calculate_available_slots()
        self.lbl_slot_capacity.config(
            text=f"Anlık RACH Slot Kapasitesi: {available_slots} Slot"
        )

        non_barred.sort(key=lambda x: x["priority"])

        # ---- Çakışma olasılığı ----
        if self.network_load < 40:
            collision_probability = 0.10
        elif self.network_load < 55:
            collision_probability = 0.25
        elif self.network_load < 70:
            collision_probability = 0.45
        elif self.network_load < 85:
            collision_probability = 0.65
        else:
            collision_probability = 0.85

        accepted  = []
        collided  = []

        if len(non_barred) <= available_slots:
            accepted = non_barred
        else:
            accepted = non_barred[:available_slots]
            overflow = non_barred[available_slots:]
            for item in overflow:
                if random.random() < collision_probability:
                    collided.append(item)
                else:
                    accepted.append(item)

        # ---- Sonuçları ata ----
        for item in barred:
            item["result"] = "BARRING ENGELİ"
            self.blocked_count += 1

        for item in collided:
            # Backoff: üstel bekleme süresi hesapla
            attempt = item.get("attempt", 1)
            if attempt < self.BACKOFF_MAX_ATTEMPTS:
                wait_ticks = self.BACKOFF_BASE ** attempt + random.randint(0, attempt)
                # Kuyruğa gidecek kopya backoff bilgisini taşır
                queued = dict(item)
                queued["attempt"] = attempt + 1
                queued["next_tick"] = self.tick_count + wait_ticks
                queued["barred"] = False
                queued["result"] = f"BACKOFF ({wait_ticks} tur)"
                self.backoff_queue.append(queued)
                self.backoff_count += 1
                # Tabloda görünen orijinal: çakışma bilgisi + kaç tur bekleyeceği
                item["result"] = f"RACH ÇAKIŞMASI → {wait_ticks} tur bekle"
            else:
                item["result"] = "TIMEOUT"
                self.timeout_count += 1
            self.collided_count += 1

        for item in accepted:
            if item["result"] == "" or item["result"].startswith("BACKOFF"):
                item["result"] = f"ERİŞİM BAŞARILI (deneme {item.get('attempt', 1)})" if item.get("attempt", 1) > 1 else "ERİŞİM BAŞARILI"
                self.success_count += 1

        all_items = attempts_this_round
        self.total_attempts += len(all_items)

        if self.total_attempts > 0:
            self.collision_rate = int((self.collided_count / self.total_attempts) * 100)
        else:
            self.collision_rate = 0

        ok_count        = sum(1 for x in all_items if x["result"] == "ERİŞİM BAŞARILI")
        barred_count    = sum(1 for x in all_items if x["result"] == "BARRING ENGELİ")
        collision_count = sum(1 for x in all_items if x["result"].startswith("RACH ÇAKIŞMASI") or x["result"] == "TIMEOUT")
        timeout_count   = sum(1 for x in all_items if x["result"] == "TIMEOUT")

        # ---- Geçmiş ----
        MAX_HISTORY = 200
        if len(self.history_tables) >= MAX_HISTORY:
            oldest_key = min(self.history_tables.keys())
            del self.history_tables[oldest_key]
            if oldest_key in self.history_stats:
                del self.history_stats[oldest_key]

        self.history_tables[self.tick_count] = [dict(item) for item in all_items]
        self.history_stats[self.tick_count] = {
            "total": len(all_items),
            "success": ok_count,
            "blocked": barred_count,
            "collided": collision_count,
            "timeout": timeout_count,
            "backoff_queue": len(self.backoff_queue),
            "collision_rate": int((collision_count / len(all_items)) * 100) if all_items else 0,
            "success_rate": int((ok_count / len(all_items)) * 100) if all_items else 0,
            "network_load": self.network_load,
            "barring_level": self.barring_level,
            "available_slots": available_slots,
            "seismic_value": self.last_seismic_value,
        }

        self.show_table_rows(all_items)
        self.lbl_table_info.config(text=f"Gösterilen Tablo: Son Tur (Tur {self.tick_count})")

        log_message = (
            f"[TUR {self.tick_count}] "
            f"{len(all_items)} deneme | "
            f"Başarılı: {ok_count} | "
            f"Engelli: {barred_count} | "
            f"Çakışma→Backoff: {collision_count - timeout_count} | "
            f"Timeout: {timeout_count} | "
            f"Kuyruk: {len(self.backoff_queue)}"
        )
        self.add_log(log_message, tick_no=self.tick_count)
        self.update_graph()
        self.refresh_priority_stats_window()

    def calculate_priority_statistics(self):
        stats = {}

        for name, data in self.user_classes.items():
            stats[name] = {
                "priority": data["priority"],
                "total": 0,
                "success": 0,
                "blocked": 0,
                "collided": 0,
                "timeout": 0,
                "rate": 0,
            }

        for rows in self.history_tables.values():
            for item in rows:
                cls = item.get("class", "")
                if cls not in stats:
                    continue

                result = item.get("result", "")
                stats[cls]["total"] += 1

                if result.startswith("ERİŞİM BAŞARILI"):
                    stats[cls]["success"] += 1
                elif result == "BARRING ENGELİ":
                    stats[cls]["blocked"] += 1
                elif result == "TIMEOUT":
                    stats[cls]["timeout"] += 1
                    stats[cls]["collided"] += 1
                elif result.startswith("RACH ÇAKIŞMASI"):
                    stats[cls]["collided"] += 1

        for cls, s in stats.items():
            total = s["total"]
            success = s["success"]
            s["rate"] = int((success / total) * 100) if total > 0 else 0

        return stats

    def open_priority_stats_window(self):
        if self.priority_stats_window is not None:
            try:
                if self.priority_stats_window.winfo_exists():
                    self.priority_stats_window.lift()
                    self.priority_stats_window.focus_force()
                    self.refresh_priority_stats_window()
                    return
            except Exception:
                pass

        win = tk.Toplevel(self.root)
        self.priority_stats_window = win
        self.priority_stats_cards = {}
        self.priority_stats_tree = None

        win.title("Canlı Öncelik İstatistikleri")
        win.geometry("920x430")
        win.configure(bg="#d9d9d9")
        win.resizable(True, True)

        def on_close():
            self.priority_stats_window = None
            self.priority_stats_cards = {}
            self.priority_stats_tree = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)

        title = tk.Label(
            win,
            text="ÖNCELİK SINIFLARINA GÖRE CANLI BAŞARI ORANLARI",
            font=("Courier New", 14, "bold"),
            bg="#808080",
            fg="white",
            pady=8
        )
        title.pack(fill="x", padx=8, pady=8)

        card_frame = tk.Frame(win, bg="#d9d9d9")
        card_frame.pack(fill="x", padx=8, pady=4)

        colors = {
            "AFAD / Ambulans / İtfaiye": "#b30000",
            "Kamu / Kritik Altyapı": "#cc5500",
            "Normal Vatandaş": "#004c99",
            "Düşük Öncelik / Toplu Trafik": "#666666",
        }

        for cls, data in sorted(self.user_classes.items(), key=lambda x: x[1]["priority"]):
            card = tk.Label(
                card_frame,
                text=f"{cls}\nBaşarı: %0 (0/0)",
                font=("Courier New", 10, "bold"),
                bg=colors.get(cls, "#666666"),
                fg="white",
                relief="raised",
                bd=4,
                padx=8,
                pady=8,
                width=28,
                justify="center"
            )
            card.pack(side="left", fill="x", expand=True, padx=4)
            self.priority_stats_cards[cls] = card

        table_frame = tk.Frame(win, bg="#d9d9d9")
        table_frame.pack(fill="both", expand=True, padx=8, pady=8)

        cols = ("sinif", "oncelik", "toplam", "basarili", "engelli", "cakisma", "timeout", "basari")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        self.priority_stats_tree = tree

        tree.heading("sinif", text="Kullanıcı Sınıfı")
        tree.heading("oncelik", text="Öncelik")
        tree.heading("toplam", text="Toplam")
        tree.heading("basarili", text="Başarılı")
        tree.heading("engelli", text="Engelli")
        tree.heading("cakisma", text="RACH/Backoff")
        tree.heading("timeout", text="Timeout")
        tree.heading("basari", text="Başarı %")

        tree.column("sinif", width=230, anchor="center")
        tree.column("oncelik", width=70, anchor="center")
        tree.column("toplam", width=70, anchor="center")
        tree.column("basarili", width=80, anchor="center")
        tree.column("engelli", width=80, anchor="center")
        tree.column("cakisma", width=100, anchor="center")
        tree.column("timeout", width=80, anchor="center")
        tree.column("basari", width=80, anchor="center")

        tree.pack(side="left", fill="both", expand=True)

        scroll = tk.Scrollbar(table_frame, command=tree.yview)
        scroll.pack(side="right", fill="y")
        tree.config(yscrollcommand=scroll.set)

        info = tk.Label(
            win,
            text="Bu pencere açıkken istatistikler her tur otomatik güncellenir.",
            font=("Courier New", 10, "bold"),
            bg="#fff2cc",
            fg="black",
            relief="groove",
            bd=2,
            padx=8,
            pady=6
        )
        info.pack(fill="x", padx=8, pady=(0, 8))

        self.refresh_priority_stats_window()

    def refresh_priority_stats_window(self):
        if self.priority_stats_window is None:
            return

        try:
            if not self.priority_stats_window.winfo_exists():
                self.priority_stats_window = None
                self.priority_stats_cards = {}
                self.priority_stats_tree = None
                return
        except Exception:
            return

        stats = self.calculate_priority_statistics()

        for cls, s in sorted(stats.items(), key=lambda x: x[1]["priority"]):
            card = self.priority_stats_cards.get(cls)
            if card is not None:
                card.config(
                    text=f"{cls}\nBaşarı: %{s['rate']} ({s['success']}/{s['total']})"
                )

        tree = self.priority_stats_tree
        if tree is not None:
            tree.delete(*tree.get_children())

            for cls, s in sorted(stats.items(), key=lambda x: x[1]["priority"]):
                tree.insert(
                    "",
                    "end",
                    values=(
                        cls,
                        f"P{s['priority']}",
                        s["total"],
                        s["success"],
                        s["blocked"],
                        s["collided"],
                        s["timeout"],
                        f"%{s['rate']}",
                    )
                )

    def calculate_total_success_rate(self):
        if self.total_attempts <= 0:
            return 0
        return int((self.success_count / self.total_attempts) * 100)

    def show_total_status(self):
        self.set_button_pressed_effect(self.btn_total_status)
        self.showing_total_status = True
        self.selected_tick_no = None
        self.lbl_selected_log.config(text="Seçilen Log: Yok / Genel toplam görünümü")
        self.lbl_table_info.config(text="Gösterilen Tablo: Son Tur")
        self.refresh_labels(self.last_seismic_value)

    def show_tick_status(self, tick_no):
        stats = self.history_stats.get(tick_no)
        if not stats:
            return
        self.showing_total_status = False
        self.lbl_seismic.config(text=f"Sismik Seviye (Tur): {stats['seismic_value']:.2f}")
        self.lbl_network.config(text=f"Şebeke Yükü (Tur): %{stats['network_load']}")
        self.lbl_barring.config(text=f"Barring Seviyesi (Tur): %{stats['barring_level']}")
        self.lbl_collision.config(text=f"RACH Çakışma (Tur): %{stats['collision_rate']}")
        self.lbl_total.config(text=f"Deneme (Tur): {stats['total']}")
        self.lbl_success.config(text=f"Başarılı (Tur): {stats['success']}")
        self.lbl_blocked.config(text=f"Barring Engeli (Tur): {stats['blocked']}")
        self.lbl_collided.config(text=f"Çakışan Paket (Tur): {stats['collided']}")
        self.lbl_success_rate.config(text=f"Tur Başarı Oranı: %{stats['success_rate']}", bg="#00c853")
        if hasattr(self, "lbl_slot_capacity"):
            self.lbl_slot_capacity.config(
                text=f"Anlık RACH Slot Kapasitesi (Tur): {stats.get('available_slots', self.calculate_available_slots())} Slot"
            )

    # =========================================================
    # LABEL
    # =========================================================
    def refresh_labels(self, seismic_value):
        if not self.showing_total_status and self.selected_tick_no in self.history_stats:
            self.show_tick_status(self.selected_tick_no)
            return

        mode_text = "AKTİF" if self.earthquake_mode else "KAPALI"
        mode_color = "#ff4d4d" if self.earthquake_mode else "#66cc66"

        pause_text = "DURDURULDU" if self.paused else "ÇALIŞIYOR"
        pause_color = "#ffcc66" if self.paused else "#99ff99"

        self.lbl_earthquake.config(text=f"Deprem Modu: {mode_text}", bg=mode_color)
        self.lbl_mode_time.config(text=f"Deprem Modu Süresi: {self.get_earthquake_mode_duration()} sn")
        self.lbl_pause.config(text=f"Simülasyon Durumu: {pause_text}", bg=pause_color)
        self.lbl_seismic.config(text=f"Sismik Seviye: {seismic_value:.2f}")
        self.lbl_network.config(text=f"Şebeke Yükü: %{self.network_load}")
        self.lbl_barring.config(text=f"Barring Seviyesi: %{self.barring_level}")
        self.lbl_collision.config(text=f"RACH Çakışma: %{self.collision_rate}")
        self.lbl_total.config(text=f"Toplam Deneme: {self.total_attempts}")
        self.lbl_success.config(text=f"Başarılı Erişim: {self.success_count}")
        self.lbl_blocked.config(text=f"Barring Engeli: {self.blocked_count}")
        self.lbl_collided.config(text=f"Çakışan Paket: {self.collided_count}")
        self.lbl_backoff.config(text=f"Backoff Kuyruğu: {len(self.backoff_queue)}")
        self.lbl_timeout.config(text=f"Timeout (Max Deneme): {self.timeout_count}")
        self.lbl_success_rate.config(text=f"Toplam Başarı Oranı: %{self.calculate_total_success_rate()}", bg="#00c853")
        if hasattr(self, "lbl_slot_capacity"):
            self.lbl_slot_capacity.config(
                text=f"Anlık RACH Slot Kapasitesi: {self.calculate_available_slots()} Slot"
            )

    # =========================================================
    # ANA DÖNGÜ
    # =========================================================
    def update_loop(self):
        if self.paused:
            self.refresh_labels(self.last_seismic_value)
            self.root.after(1000, self.update_loop)
            return

        seismic_value = self.get_seismic_value()
        self.last_seismic_value = seismic_value

        self.check_earthquake_trigger(seismic_value)
        self.update_network_conditions()

        # Manuel trafik butonlarından gelen yük sadece bu turda uygulanır.
        # Böylece butona basmak yeni tur oluşturmaz; mevcut zaman akışındaki sıradaki tur yoğunlaşır.
        next_round_boost = self.pending_traffic_boost
        next_round_label = self.pending_traffic_label
        self.pending_traffic_boost = 0
        self.pending_traffic_label = None

        if self.earthquake_mode:
            total_boost = 25 + next_round_boost
        else:
            total_boost = next_round_boost

        if next_round_boost > 0:
            self.add_log(f"{next_round_label} trafik yükü bu turda uygulandı. Ek yük: {next_round_boost}")

        self.generate_traffic(manual_boost=total_boost)

        self.refresh_labels(seismic_value)
        self.root.after(1000, self.update_loop)

if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style()
    style.theme_use("default")
    app = RachSimulationApp(root)
    root.mainloop()
